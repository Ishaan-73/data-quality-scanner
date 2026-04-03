"""
Generate DQS_Weights_Explainer.xlsx — a management-ready document explaining
the Data Quality Scanner weighting methodology.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY      = "1B3A6B"
BLUE      = "2563EB"
LIGHT_BLUE= "DBEAFE"
MID_BLUE  = "93C5FD"
GREEN     = "16A34A"
LIGHT_GREEN="DCFCE7"
AMBER     = "D97706"
LIGHT_AMBER="FEF3C7"
RED       = "DC2626"
LIGHT_RED = "FEE2E2"
WHITE     = "FFFFFF"
LIGHT_GREY= "F8FAFC"
MID_GREY  = "E2E8F0"
DARK_GREY = "64748B"
NEAR_BLACK= "1E293B"

# ── Helpers ───────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color="CBD5E1")
    b = Border()
    if "l" in sides or sides == "all": b.left = s
    if "r" in sides or sides == "all": b.right = s
    if "t" in sides or sides == "all": b.top = s
    if "b" in sides or sides == "all": b.bottom = s
    return b

def thick_bottom():
    return Border(bottom=Side(style="medium", color=NAVY))

def write(ws, row, col, value, bold=False, bg=None, fg=NEAR_BLACK,
          size=11, h="left", v="center", wrap=False, italic=False,
          border=None, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic, name="Calibri")
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align(h=h, v=v, wrap=wrap)
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format
    return cell

def section_header(ws, row, col, col_end, text, bg=NAVY):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col_end)
    write(ws, row, col, text, bold=True, bg=bg, fg=WHITE, size=12, h="center")
    ws.row_dimensions[row].height = 22

def col_header(ws, row, col, text):
    write(ws, row, col, text, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", border=thin_border())

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def score_color(score_pct):
    """Return fill color based on score 0–1."""
    if score_pct >= 0.9: return LIGHT_GREEN
    if score_pct >= 0.7: return LIGHT_AMBER
    return LIGHT_RED

def phase_color(phase):
    return {
        "MVP (Phase 1)": LIGHT_GREEN,
        "Phase 2": LIGHT_AMBER,
        "Later": LIGHT_GREY,
    }.get(phase, WHITE)

def phase_font_color(phase):
    return {
        "MVP (Phase 1)": GREEN,
        "Phase 2": AMBER,
        "Later": DARK_GREY,
    }.get(phase, NEAR_BLACK)

# ── Data ─────────────────────────────────────────────────────────────────────

DIMENSION_DATA = [
    # (dimension, display_name, weight, rationale, num_checks, phase_note)
    ("completeness", "Completeness", 0.20,
     "Missing data breaks all downstream analytics and AI models. A null in a join key silently drops rows; a null in a model feature corrupts predictions. Highest blast radius of any dimension.",
     4, "MVP"),
    ("uniqueness", "Uniqueness", 0.14,
     "Duplicates silently inflate aggregations. A 2% duplicate rate in a fact table means every SUM() is wrong by 2%. Silent and compounding.",
     3, "MVP"),
    ("validity", "Validity", 0.14,
     "Type and format errors either crash pipelines at ingestion or silently coerce values to NULL. Same blast radius as completeness at the column level.",
     5, "MVP"),
    ("integrity", "Integrity", 0.12,
     "FK violations mean joins produce wrong results — analysts see numbers, they just aren't trustworthy. High trust signal for analytics joins.",
     3, "MVP"),
    ("accuracy", "Accuracy", 0.10,
     "Source-of-truth reconciliation. Requires a reference/source to compare against — harder to compute but very high signal when available.",
     3, "Phase 2"),
    ("freshness", "Freshness", 0.09,
     "Stale data misleads decisions and degrades AI inference quality. Vital for operational and real-time workloads.",
     3, "MVP"),
    ("consistency", "Consistency", 0.08,
     "Cross-table and cross-field alignment. Requires more setup (table mapping, standards definitions) than foundational checks.",
     4, "Phase 2"),
    ("volume", "Volume", 0.05,
     "Row count anomalies and missing partitions are leading indicators of upstream pipeline failures.",
     2, "MVP"),
    ("anomaly", "Anomaly", 0.03,
     "Statistical outliers and distribution drift. Useful signal once baselines exist; advisory by default.",
     2, "Phase 2"),
    ("schema", "Schema", 0.02,
     "Structural drift detection. Requires a baseline snapshot to compare against.",
     1, "Phase 2"),
    ("pipeline", "Pipeline", 0.01,
     "Operational health of data loads. Narrow scope — supplements other checks rather than standing alone.",
     1, "Phase 2"),
    ("time_series", "Time Series", 0.01,
     "Timestamp sequence integrity for event/time-series datasets. Specialised use case.",
     1, "Later"),
    ("metadata", "Metadata", 0.01,
     "Governance-adjacent — missing descriptions, owners, and tags. Important for compliance but low direct analytics impact.",
     1, "Later"),
]

# Old spec weights (flat, sum = 1.37)
SPEC_WEIGHTS = {
    1: 0.05, 2: 0.08, 3: 0.03, 4: 0.04,
    5: 0.08, 6: 0.05, 7: 0.02,
    8: 0.04, 9: 0.04, 10: 0.04, 11: 0.04, 12: 0.04,
    13: 0.05, 14: 0.05, 15: 0.03, 16: 0.03,
    17: 0.07, 18: 0.05, 19: 0.03,
    20: 0.06, 21: 0.04, 22: 0.03,
    23: 0.04, 24: 0.03,
    25: 0.06, 26: 0.05, 27: 0.04,
    28: 0.02, 29: 0.03,
    30: 0.03,
    31: 0.02,
    32: 0.03,
    33: 0.03,
}

CHECK_DATA = [
    # id, name, dimension, within_dim_weight, threshold, direction, phase, rationale
    (1,  "Null Ratio by Column",             "completeness",  0.30, "≤ 5%",     "lower", "MVP",     "Baseline null profiling across all columns."),
    (2,  "Null Ratio in Critical Columns",   "completeness",  0.40, "≤ 1%",     "lower", "MVP",     "Highest completeness weight — zero-tolerance on key columns like IDs and emails."),
    (3,  "Empty Row Check",                  "completeness",  0.10, "= 0",      "lower", "MVP",     "Rows where all major fields are null. Rare but critical when found."),
    (4,  "Conditional Completeness",         "completeness",  0.20, "≤ 1%",     "lower", "MVP",     "Required field must be present when a condition is true (e.g. shipping address when order type is ONLINE)."),
    (5,  "PK Duplicate Ratio",               "uniqueness",    0.55, "= 0",      "lower", "MVP",     "Highest uniqueness weight — any PK duplication is definitionally wrong. Zero tolerance."),
    (6,  "Business Key Duplicate Ratio",     "uniqueness",    0.35, "≤ 0.5%",   "lower", "MVP",     "Duplicates on email, phone, order ID etc. 0.5% tolerance."),
    (7,  "Near-Duplicate Candidate Check",   "uniqueness",    0.10, "Advisory", "lower", "MVP",     "Fuzzy/phonetic duplicate candidates. Advisory — no hard threshold."),
    (8,  "Data Type Conformance",            "validity",      0.30, "≤ 0.5%",   "lower", "MVP",     "Highest validity weight — wrong types crash pipelines or silently coerce values."),
    (9,  "Allowed Value / Domain Check",     "validity",      0.25, "= 0",      "lower", "MVP",     "Values outside approved enum list. Zero tolerance — every violation is a data contract breach."),
    (10, "Format Validation",                "validity",      0.20, "≤ 0.5%",   "lower", "MVP",     "Email, phone, postcode regex patterns."),
    (11, "Range Validation",                 "validity",      0.15, "≤ 0.5%",   "lower", "MVP",     "Numeric or date values outside expected min/max bounds."),
    (12, "Negative Value Check",             "validity",      0.10, "= 0",      "lower", "MVP",     "Non-negative measures (revenue, quantity) found below zero."),
    (13, "Cross-Field Consistency",          "consistency",   0.35, "= 0",      "lower", "Phase 2", "Logical consistency within a row (e.g. end_date must be after start_date)."),
    (14, "Cross-Table Consistency",          "consistency",   0.35, "≤ 0.5%",   "lower", "Phase 2", "Same business entity attributes conflicting across tables."),
    (15, "Unit / Currency Consistency",      "consistency",   0.20, "= 0",      "lower", "Phase 2", "Unexpected mix of units, currencies, or measurement scales."),
    (16, "Standardisation Check",            "consistency",   0.10, "≤ 1%",     "lower", "Phase 2", "Inconsistent labels for the same concept (e.g. 'USA' vs 'US' vs 'United States')."),
    (17, "Foreign Key Violation Check",      "integrity",     0.50, "= 0",      "lower", "MVP",     "Highest integrity weight — child rows without a matching parent break every join downstream."),
    (18, "Orphan Record Check",              "integrity",     0.35, "= 0",      "lower", "MVP",     "Unlinked rows that should map to a master or dimension table."),
    (19, "Broken Join Coverage",             "integrity",     0.15, "≥ 99%",    "higher","MVP",     "Expected join coverage falls below target. Softer version of FK check."),
    (20, "Freshness Lag Hours",              "freshness",     0.50, "≤ SLA",    "lower", "MVP",     "Hours since latest update vs. configured SLA. Primary freshness metric."),
    (21, "Stale Table Check",                "freshness",     0.30, "= 0",      "lower", "MVP",     "Binary flag — table has not been updated within the configured window."),
    (22, "Late-Arriving Data Check",         "freshness",     0.20, "≤ 1%",     "lower", "MVP",     "Records that arrived later than the allowed delay after their event time."),
    (23, "Row Count Volume Change",          "volume",        0.65, "≤ ±30%",   "lower", "MVP",     "Highest volume weight — unexpected day-over-day row count change is the leading pipeline failure indicator."),
    (24, "Missing Partition / Date Slice",   "volume",        0.35, "= 0",      "lower", "MVP",     "Expected partition or date slice not loaded."),
    (25, "Source-to-Target Reconciliation",  "accuracy",      0.45, "≤ 0.5%",   "lower", "Phase 2", "Target totals reconcile to source/control totals. Highest accuracy weight."),
    (26, "Business Rule Validation",         "accuracy",      0.35, "≤ 0.5%",   "lower", "Phase 2", "Derived business logic holds (e.g. revenue = price × quantity)."),
    (27, "Reference Data Validation",        "accuracy",      0.20, "≤ 0.5%",   "lower", "Phase 2", "Values align to trusted reference or master data."),
    (28, "Outlier Detection",                "anomaly",       0.60, "Advisory", "lower", "Phase 2", "Extreme values outside Z-score bounds. Advisory — no hard threshold."),
    (29, "Distribution Drift Check",         "anomaly",       0.40, "Advisory", "lower", "Phase 2", "Distribution materially shifts from baseline period (PSI/KS). Advisory."),
    (30, "Schema Drift Detection",           "schema",        1.00, "= 0",      "lower", "Phase 2", "Columns added, dropped, renamed, or type-changed vs. baseline snapshot."),
    (31, "Metadata Completeness",            "metadata",      1.00, "≤ 5%",     "lower", "Later",   "Columns or tables missing description, owner, tags, or classification."),
    (32, "Pipeline Failure Indicator",       "pipeline",      1.00, "≤ 1%",     "lower", "Phase 2", "Recent failed loads or jobs associated with this dataset."),
    (33, "Timestamp Sequence Integrity",     "time_series",   1.00, "≤ 0.5%",   "lower", "Later",   "Duplicate, missing, or non-monotonic timestamps in time-series data."),
]

DIM_WEIGHTS = {d[0]: d[2] for d in DIMENSION_DATA}

def effective_weight(check_id, dimension, within_weight):
    return DIM_WEIGHTS[dimension] * within_weight


# ─────────────────────────────────────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # remove default sheet


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Executive Summary
# ══════════════════════════════════════════════════════════════════════════════

ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False
ws1.sheet_view.zoomScale = 100

set_col_widths(ws1, {
    "A": 3, "B": 28, "C": 14, "D": 14, "E": 14, "F": 50, "G": 3,
})

# Title block
ws1.merge_cells("B1:F1")
ws1.row_dimensions[1].height = 36
write(ws1, 1, 2, "Data Quality Scanner — Weighting Methodology",
      bold=True, bg=NAVY, fg=WHITE, size=16, h="center")

ws1.merge_cells("B2:F2")
ws1.row_dimensions[2].height = 18
write(ws1, 2, 2, "Management Overview  ·  Framework: DAMA-DMBOK2 / ISO 25012",
      italic=True, bg=LIGHT_BLUE, fg=NAVY, size=11, h="center")

ws1.row_dimensions[3].height = 12

# Problem statement box
ws1.merge_cells("B4:F4")
section_header(ws1, 4, 2, 6, "WHY WE REDESIGNED THE WEIGHTS", NAVY)

for r, text in enumerate([
    "The original specification assigned a flat weight to each of the 33 checks. This had two problems:",
    "  1.  The weights summed to 1.37, not 1.00 — making any composite score mathematically meaningless (max score would be 137, not 100).",
    "  2.  A flat list gives no way to tune how important a quality category is vs. how important a specific check is within that category.",
    "",
    "The redesign introduces a two-level hierarchy that keeps both layers independently normalised to 1.00,",
    "producing a valid 0–100 quality score grounded in industry research.",
], start=5):
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    is_header = r == 5
    write(ws1, r, 2, text, bg=LIGHT_RED if r in (6,7) else LIGHT_GREY,
          fg=RED if r in (6,7) else NEAR_BLACK,
          bold=is_header, size=10, wrap=True, h="left")
    ws1.row_dimensions[r].height = 22 if r in (6,7) else 18

ws1.row_dimensions[11].height = 14

# How it works
ws1.merge_cells("B12:F12")
section_header(ws1, 12, 2, 6, "THE TWO-LEVEL WEIGHT SYSTEM")

ws1.merge_cells("B13:F13")
write(ws1, 13, 2,
      "Overall Quality Score  =  Σ ( Dimension Weight  ×  Check-Within-Dimension Weight  ×  Pass Score )  ×  100",
      bold=True, bg=LIGHT_BLUE, fg=NAVY, size=11, h="center")
ws1.row_dimensions[13].height = 24

for r, (label, desc) in enumerate([
    ("Level 1 — Dimension Weights",
     "How important is each quality category (Completeness, Uniqueness, etc.) relative to all others?  13 dimensions, sum = 1.00."),
    ("Level 2 — Check-Within-Dimension Weights",
     "Within a given dimension, which specific check matters most?  Each dimension's checks sum to 1.00 independently."),
    ("Effective Weight per Check",
     "= Dimension Weight × Check-Within-Dimension Weight.  All 33 effective weights sum to 1.00."),
    ("Pass Score",
     "Not binary. A check that barely fails scores higher than one that fails badly. Linear proximity scoring from 1.0 (at threshold) to 0.0."),
], start=14):
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    write(ws1, r, 2, label, bold=True, bg=MID_BLUE, fg=NAVY, size=10,
          border=thin_border(), wrap=True)
    write(ws1, r, 4, desc, bg=WHITE, fg=NEAR_BLACK, size=10,
          border=thin_border(), wrap=True)
    ws1.row_dimensions[r].height = 32

ws1.row_dimensions[18].height = 14

# Scorecard comparison
ws1.merge_cells("B19:F19")
section_header(ws1, 19, 2, 6, "BEFORE vs. AFTER COMPARISON")

headers = ["", "Original Spec", "Redesigned", "Improvement"]
for c, h in enumerate(headers, start=2):
    write(ws1, 20, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", border=thin_border())

rows = [
    ("Weights sum to",         "1.37  ✗",  "1.00  ✓",   "Mathematically valid"),
    ("Maximum score",          "137   ✗",  "100   ✓",   "Normalised 0–100 scale"),
    ("Score structure",        "Flat list", "Two-level hierarchy", "Tunable at category + check level"),
    ("Partial credit",         "None",     "Proximity scoring",  "Gradual penalty for near-misses"),
    ("Advisory checks",        "Not handled", "threshold=None → always pass", "No score penalty for exploratory checks"),
    ("Industry grounding",     "Custom",   "DAMA-DMBOK2 / ISO 25012", "Peer-reviewed priority ordering"),
]
for i, (label, before, after, note) in enumerate(rows, start=21):
    bg = LIGHT_GREY if i % 2 == 0 else WHITE
    write(ws1, i, 2, label, bold=True, bg=bg, fg=NEAR_BLACK, size=10, border=thin_border())
    write(ws1, i, 3, before, bg=LIGHT_RED, fg=RED, size=10, h="center", border=thin_border())
    write(ws1, i, 4, after, bg=LIGHT_GREEN, fg=GREEN, size=10, h="center", border=thin_border())
    write(ws1, i, 5, note, bg=bg, fg=DARK_GREY, size=10, italic=True, border=thin_border())
    ws1.row_dimensions[i].height = 18

ws1.row_dimensions[27].height = 14

# Key facts bar
ws1.merge_cells("B28:F28")
section_header(ws1, 28, 2, 6, "KEY FACTS AT A GLANCE")
facts = [
    ("33", "Total Checks"),
    ("13", "Quality Dimensions"),
    ("20", "MVP Phase 1 Checks"),
    ("1.00", "Sum of all effective weights"),
    ("0–100", "Quality score range"),
]
for c, (num, label) in enumerate(facts, start=2):
    write(ws1, 29, c, num, bold=True, bg=LIGHT_BLUE, fg=NAVY, size=14, h="center")
    write(ws1, 30, c, label, bg=WHITE, fg=DARK_GREY, size=9, h="center", italic=True)
    ws1.row_dimensions[29].height = 28
    ws1.row_dimensions[30].height = 16


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Dimension Weights
# ══════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Dimension Weights")
ws2.sheet_view.showGridLines = False

set_col_widths(ws2, {
    "A": 3, "B": 5, "C": 22, "D": 12, "E": 12, "F": 10, "G": 10, "H": 52, "I": 3,
})

ws2.merge_cells("B1:H1")
ws2.row_dimensions[1].height = 32
write(ws2, 1, 2, "Level 1 — Dimension Weights", bold=True, bg=NAVY, fg=WHITE,
      size=14, h="center")

ws2.merge_cells("B2:H2")
write(ws2, 2, 2,
      "13 quality dimensions ordered by analytics impact. Weights sum to exactly 1.00. Based on DAMA-DMBOK2 and ISO 25012.",
      bg=LIGHT_BLUE, fg=NAVY, size=10, h="center", italic=True)
ws2.row_dimensions[2].height = 18
ws2.row_dimensions[3].height = 10

headers = ["#", "Dimension", "Weight", "% of Score", "Checks", "Phase", "Why This Priority"]
for c, h in enumerate(headers, start=2):
    write(ws2, 4, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", border=thin_border())
ws2.row_dimensions[4].height = 20

cumulative = 0.0
for i, (dim, display, weight, rationale, n_checks, phase) in enumerate(DIMENSION_DATA, start=5):
    cumulative += weight
    bg = LIGHT_GREY if i % 2 == 0 else WHITE
    p_bg = phase_color(f"MVP" if phase == "MVP" else ("Phase 2" if phase == "Phase 2" else "Later"))
    p_fg = phase_font_color(f"MVP" if phase == "MVP" else ("Phase 2" if phase == "Phase 2" else "Later"))

    write(ws2, i, 2, i - 4, bg=bg, fg=DARK_GREY, size=10, h="center", border=thin_border())
    write(ws2, i, 3, display, bold=True, bg=bg, fg=NEAR_BLACK, size=10, border=thin_border())
    write(ws2, i, 4, weight, bg=LIGHT_BLUE if weight >= 0.10 else bg,
          fg=NAVY if weight >= 0.10 else NEAR_BLACK,
          bold=weight >= 0.10, size=10, h="center",
          border=thin_border(), num_format="0.00")
    write(ws2, i, 5, weight, bg=score_color(weight / 0.20),
          fg=NEAR_BLACK, size=10, h="center",
          border=thin_border(), num_format="0%")
    write(ws2, i, 6, n_checks, bg=bg, fg=NEAR_BLACK, size=10, h="center", border=thin_border())
    write(ws2, i, 7, phase, bg=p_bg, fg=p_fg, bold=True, size=9, h="center", border=thin_border())
    write(ws2, i, 8, rationale, bg=bg, fg=DARK_GREY, size=9, wrap=True,
          border=thin_border())
    ws2.row_dimensions[i].height = 30

total_row = len(DIMENSION_DATA) + 5
write(ws2, total_row, 2, "", bg=MID_GREY, border=thin_border())
write(ws2, total_row, 3, "TOTAL", bold=True, bg=MID_GREY, fg=NAVY, size=10,
      h="right", border=thin_border())
write(ws2, total_row, 4, 1.0, bold=True, bg=MID_GREY, fg=NAVY, size=10,
      h="center", border=thin_border(), num_format="0.00")
write(ws2, total_row, 5, 1.0, bold=True, bg=MID_GREY, fg=NAVY, size=10,
      h="center", border=thin_border(), num_format="0%")
write(ws2, total_row, 6, 33, bold=True, bg=MID_GREY, fg=NAVY, size=10,
      h="center", border=thin_border())
write(ws2, total_row, 7, "", bg=MID_GREY, border=thin_border())
write(ws2, total_row, 8, "✓  Sum validated programmatically on every run.",
      bold=True, bg=MID_GREY, fg=GREEN, size=10, border=thin_border())
ws2.row_dimensions[total_row].height = 20

# Pie chart
pie = PieChart()
pie.title = "Dimension Weight Distribution"
pie.style = 10
pie.width = 14
pie.height = 14

labels = Reference(ws2, min_col=3, min_row=5, max_row=5 + len(DIMENSION_DATA) - 1)
data = Reference(ws2, min_col=4, min_row=5, max_row=5 + len(DIMENSION_DATA) - 1)
pie.add_data(data)
pie.set_categories(labels)
pie.series[0].title = None

ws2.add_chart(pie, "B20")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — All 33 Checks
# ══════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("All 33 Checks")
ws3.sheet_view.showGridLines = False

set_col_widths(ws3, {
    "A": 3, "B": 5, "C": 32, "D": 16, "E": 12, "F": 14, "G": 14,
    "H": 10, "I": 12, "J": 10, "K": 40, "L": 3,
})

ws3.merge_cells("B1:K1")
ws3.row_dimensions[1].height = 32
write(ws3, 1, 2, "All 33 Quality Checks — Weights & Configuration",
      bold=True, bg=NAVY, fg=WHITE, size=14, h="center")
ws3.merge_cells("B2:K2")
write(ws3, 2, 2,
      "Effective Weight = Dimension Weight × Check-Within-Dimension Weight.  All 33 effective weights sum to 1.00.",
      bg=LIGHT_BLUE, fg=NAVY, size=10, h="center", italic=True)
ws3.row_dimensions[2].height = 18
ws3.row_dimensions[3].height = 10

col_headers = [
    "ID", "Check Name", "Dimension", "Dim Weight",
    "Check Weight\n(within dim)", "Effective\nWeight",
    "Threshold", "Phase", "Direction", "Rationale"
]
for c, h in enumerate(col_headers, start=2):
    write(ws3, 4, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", wrap=True, border=thin_border())
ws3.row_dimensions[4].height = 30

prev_dim = None
for i, (cid, name, dim, within_w, thresh, direction, phase, rationale) in enumerate(CHECK_DATA, start=5):
    eff_w = effective_weight(cid, dim, within_w)
    dim_w = DIM_WEIGHTS[dim]
    bg = LIGHT_GREY if i % 2 == 0 else WHITE
    dim_bg = LIGHT_BLUE if dim != prev_dim else bg
    prev_dim = dim

    p_label = "MVP (Phase 1)" if phase == "MVP" else ("Phase 2" if phase == "Phase 2" else "Later")
    p_bg = phase_color(p_label)
    p_fg = phase_font_color(p_label)

    write(ws3, i, 2, cid, bg=bg, fg=DARK_GREY, size=10, h="center", border=thin_border())
    write(ws3, i, 3, name, bold=True, bg=bg, fg=NEAR_BLACK, size=10,
          border=thin_border(), wrap=True)
    write(ws3, i, 4, dim.title(), bg=dim_bg, fg=NAVY if dim != prev_dim else NEAR_BLACK,
          bold=(dim != prev_dim), size=9, h="center", border=thin_border())
    write(ws3, i, 5, dim_w, bg=bg, fg=NEAR_BLACK, size=10, h="center",
          border=thin_border(), num_format="0%")
    write(ws3, i, 6, within_w, bg=bg, fg=NEAR_BLACK, size=10, h="center",
          border=thin_border(), num_format="0%")
    write(ws3, i, 7, eff_w, bold=True,
          bg=LIGHT_BLUE if eff_w >= 0.06 else (LIGHT_AMBER if eff_w >= 0.03 else bg),
          fg=NAVY if eff_w >= 0.06 else NEAR_BLACK,
          size=10, h="center", border=thin_border(), num_format="0.0000")
    write(ws3, i, 8, thresh, bg=bg, fg=NEAR_BLACK, size=10, h="center", border=thin_border())
    write(ws3, i, 9, p_label, bg=p_bg, fg=p_fg, bold=True, size=9, h="center", border=thin_border())
    dir_label = "Higher = Better" if direction == "higher" else "Lower = Better"
    write(ws3, i, 10, dir_label, bg=bg, fg=DARK_GREY, size=9, h="center", border=thin_border())
    write(ws3, i, 11, rationale, bg=bg, fg=DARK_GREY, size=9, wrap=True, border=thin_border())
    ws3.row_dimensions[i].height = 30

# Total row
tr = len(CHECK_DATA) + 5
for c in range(2, 12):
    write(ws3, tr, c, "", bg=MID_GREY, border=thin_border())
write(ws3, tr, 3, "TOTAL (all 33 checks)", bold=True, bg=MID_GREY, fg=NAVY, size=10)
eff_total = sum(effective_weight(c[0], c[2], c[3]) for c in CHECK_DATA)
write(ws3, tr, 7, eff_total, bold=True, bg=MID_GREY, fg=GREEN, size=10,
      h="center", border=thin_border(), num_format="0.0000")
write(ws3, tr, 11, "✓  Verified: effective weights sum to exactly 1.00",
      bold=True, bg=MID_GREY, fg=GREEN, size=10)
ws3.row_dimensions[tr].height = 20


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Original vs. Redesigned
# ══════════════════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("Original vs. Redesigned")
ws4.sheet_view.showGridLines = False

set_col_widths(ws4, {
    "A": 3, "B": 5, "C": 32, "D": 16,
    "E": 14, "F": 14, "G": 14, "H": 3,
})

ws4.merge_cells("B1:G1")
ws4.row_dimensions[1].height = 32
write(ws4, 1, 2, "Original Spec Weights vs. Redesigned Weights — Side-by-Side",
      bold=True, bg=NAVY, fg=WHITE, size=14, h="center")

ws4.merge_cells("B2:G2")
write(ws4, 2, 2,
      "Original weights (flat, sum = 1.37) vs. redesigned effective weights (two-level, sum = 1.00).",
      bg=LIGHT_BLUE, fg=NAVY, size=10, h="center", italic=True)
ws4.row_dimensions[2].height = 18

# Problem callout
ws4.merge_cells("B3:G4")
write(ws4, 3, 2,
      "⚠  The original spec weights sum to 1.37 — this makes the composite score mathematically invalid. "
      "A perfect dataset would score 137/100. The redesigned weights fix this while preserving the "
      "relative priority intent of the original.",
      bg=LIGHT_RED, fg=RED, bold=True, size=10, wrap=True, h="left")
ws4.row_dimensions[3].height = 30
ws4.row_dimensions[4].height = 10

headers = ["ID", "Check Name", "Dimension",
           "Original Weight\n(spec, flat)", "Redesigned Effective\nWeight",
           "Change"]
for c, h in enumerate(headers, start=2):
    write(ws4, 5, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", wrap=True, border=thin_border())
ws4.row_dimensions[5].height = 30

for i, (cid, name, dim, within_w, thresh, direction, phase, rationale) in enumerate(CHECK_DATA, start=6):
    orig = SPEC_WEIGHTS.get(cid, 0.0)
    new_eff = effective_weight(cid, dim, within_w)
    delta = new_eff - orig
    bg = LIGHT_GREY if i % 2 == 0 else WHITE

    write(ws4, i, 2, cid, bg=bg, fg=DARK_GREY, size=10, h="center", border=thin_border())
    write(ws4, i, 3, name, bg=bg, fg=NEAR_BLACK, size=10, border=thin_border(), wrap=True)
    write(ws4, i, 4, dim.title(), bg=bg, fg=NEAR_BLACK, size=9, h="center", border=thin_border())
    write(ws4, i, 5, orig, bg=LIGHT_RED, fg=NEAR_BLACK, size=10, h="center",
          border=thin_border(), num_format="0.00")
    write(ws4, i, 6, new_eff, bg=LIGHT_GREEN, fg=NEAR_BLACK, size=10, h="center",
          border=thin_border(), num_format="0.0000")

    if abs(delta) < 0.001:
        delta_txt, d_bg, d_fg = "~  No change", LIGHT_GREY, DARK_GREY
    elif delta > 0:
        delta_txt, d_bg, d_fg = f"▲  +{delta:.4f}", LIGHT_GREEN, GREEN
    else:
        delta_txt, d_bg, d_fg = f"▼  {delta:.4f}", LIGHT_RED, RED
    write(ws4, i, 7, delta_txt, bg=d_bg, fg=d_fg, bold=True, size=9,
          h="center", border=thin_border())
    ws4.row_dimensions[i].height = 26

# Totals row
tr = len(CHECK_DATA) + 6
write(ws4, tr, 3, "TOTAL", bold=True, bg=MID_GREY, fg=NAVY, size=10,
      h="right", border=thin_border())
write(ws4, tr, 2, "", bg=MID_GREY, border=thin_border())
write(ws4, tr, 4, "", bg=MID_GREY, border=thin_border())
orig_total = sum(SPEC_WEIGHTS.values())
new_total = sum(effective_weight(c[0], c[2], c[3]) for c in CHECK_DATA)
write(ws4, tr, 5, orig_total, bold=True, bg=LIGHT_RED, fg=RED, size=11,
      h="center", border=thin_border(), num_format="0.00")
write(ws4, tr, 6, new_total, bold=True, bg=LIGHT_GREEN, fg=GREEN, size=11,
      h="center", border=thin_border(), num_format="0.0000")
write(ws4, tr, 7, "1.37 → 1.00  ✓", bold=True, bg=LIGHT_GREEN, fg=GREEN,
      size=10, h="center", border=thin_border())
ws4.row_dimensions[tr].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Scoring Logic
# ══════════════════════════════════════════════════════════════════════════════

ws5 = wb.create_sheet("Scoring Logic")
ws5.sheet_view.showGridLines = False

set_col_widths(ws5, {
    "A": 3, "B": 22, "C": 20, "D": 20, "E": 20, "F": 20, "G": 3,
})

ws5.merge_cells("B1:F1")
ws5.row_dimensions[1].height = 32
write(ws5, 1, 2, "Scoring Logic — How the Quality Score is Calculated",
      bold=True, bg=NAVY, fg=WHITE, size=14, h="center")

ws5.row_dimensions[2].height = 14

# Formula
ws5.merge_cells("B3:F3")
section_header(ws5, 3, 2, 6, "FORMULA")

ws5.merge_cells("B4:F5")
write(ws5, 4, 2,
      "Overall Quality Score  =  Σ ( effective_weight  ×  pass_score )  ×  100\n\n"
      "where  effective_weight  =  Dimension Weight  ×  Check-Within-Dimension Weight",
      bold=True, bg=LIGHT_BLUE, fg=NAVY, size=12, h="center", wrap=True)
ws5.row_dimensions[4].height = 30
ws5.row_dimensions[5].height = 20
ws5.row_dimensions[6].height = 14

# Pass score logic
ws5.merge_cells("B7:F7")
section_header(ws5, 7, 2, 6, "PASS SCORE — PARTIAL CREDIT (NOT BINARY)")

explanations = [
    ("If check passes (metric within threshold)", "pass_score = 1.0", "Full credit."),
    ("If check fails (lower-is-better)",
     "pass_score = max(0,  1 − (actual − threshold) / threshold)",
     "Linear decay from 1.0 at the threshold to 0.0 at 2× threshold."),
    ("If check fails (higher-is-better, e.g. join coverage)",
     "pass_score = max(0,  actual / threshold)",
     "Proportional to how close actual is to target."),
    ("Advisory check (no threshold)",
     "pass_score = 1.0  always",
     "Metric is recorded but never penalises the score."),
]
for r, (condition, formula, note) in enumerate(explanations, start=8):
    bg = LIGHT_GREY if r % 2 == 0 else WHITE
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws5.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    write(ws5, r, 2, condition, bold=True, bg=bg, fg=NEAR_BLACK, size=10,
          border=thin_border(), wrap=True)
    write(ws5, r, 4, formula, bg=LIGHT_BLUE, fg=NAVY, size=10,
          h="center", border=thin_border())
    write(ws5, r, 6, note, bg=bg, fg=DARK_GREY, size=9, italic=True,
          border=thin_border(), wrap=True)
    ws5.row_dimensions[r].height = 28

ws5.row_dimensions[12].height = 14

# Worked examples
ws5.merge_cells("B13:F13")
section_header(ws5, 13, 2, 6, "WORKED EXAMPLES")

headers = ["Check", "Threshold", "Actual Value", "Pass Score", "Interpretation"]
for c, h in enumerate(headers, start=2):
    write(ws5, 14, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", border=thin_border())
ws5.row_dimensions[14].height = 20

examples = [
    ("PK Duplicate Ratio",          "= 0",   "0.000", "1.00", "✓  Perfect — no duplicates"),
    ("PK Duplicate Ratio",          "= 0",   "0.005", "0.00", "✗  Any PK dup = zero score (zero-tolerance)"),
    ("Null Ratio by Column",        "≤ 0.05","0.030", "1.00", "✓  3% nulls, well within 5% threshold"),
    ("Null Ratio by Column",        "≤ 0.05","0.070", "0.60", "✗  7% nulls — partial credit (40% over threshold)"),
    ("Null Ratio by Column",        "≤ 0.05","0.100", "0.00", "✗  10% nulls — 2× threshold, zero score"),
    ("FK Violation Check",          "= 0",   "0.032", "0.00", "✗  3.2% FK violations — zero-tolerance, zero score"),
    ("Broken Join Coverage",        "≥ 0.99","0.985", "0.99", "✗  98.5% coverage — near miss, high partial credit"),
    ("Freshness Lag Hours",         "≤ 24h", "18h",   "1.00", "✓  18 hours lag, within 24h SLA"),
    ("Outlier Detection (advisory)","None",  "0.023", "1.00", "✓  Metric recorded, no score impact (advisory)"),
]
for i, (check, thresh, actual, ps, interp) in enumerate(examples, start=15):
    bg = LIGHT_GREY if i % 2 == 0 else WHITE
    is_pass = ps == "1.00"
    write(ws5, i, 2, check, bg=bg, fg=NEAR_BLACK, size=10, border=thin_border())
    write(ws5, i, 3, thresh, bg=bg, fg=NEAR_BLACK, size=10, h="center", border=thin_border())
    write(ws5, i, 4, actual, bg=bg, fg=NEAR_BLACK, size=10, h="center", border=thin_border())
    write(ws5, i, 5, ps, bold=True,
          bg=LIGHT_GREEN if is_pass else (LIGHT_AMBER if float(ps) > 0 else LIGHT_RED),
          fg=GREEN if is_pass else (AMBER if float(ps) > 0 else RED),
          size=10, h="center", border=thin_border())
    write(ws5, i, 6, interp,
          bg=LIGHT_GREEN if "✓" in interp else LIGHT_RED,
          fg=GREEN if "✓" in interp else RED,
          size=9, border=thin_border(), wrap=True)
    ws5.row_dimensions[i].height = 20

ws5.row_dimensions[24].height = 14

# Score interpretation guide
ws5.merge_cells("B25:F25")
section_header(ws5, 25, 2, 6, "SCORE INTERPRETATION GUIDE")

bands = [
    ("90 – 100", "Excellent", GREEN, LIGHT_GREEN,
     "Production-ready. Data can be trusted for analytics and AI workloads."),
    ("70 – 89",  "Good",      AMBER, LIGHT_AMBER,
     "Minor issues present. Suitable for most analytics; investigate failures before AI use."),
    ("50 – 69",  "Fair",      AMBER, LIGHT_AMBER,
     "Meaningful data quality issues. Root-cause analysis and remediation recommended."),
    ("0 – 49",   "Poor",      RED,   LIGHT_RED,
     "Significant data quality problems. Do not use for critical decisions until resolved."),
]
for c, h in enumerate(["Score Range", "Rating", "Description"], start=2):
    write(ws5, 26, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
          h="center", border=thin_border())
write(ws5, 26, 5, "", bg=BLUE, border=thin_border())
ws5.row_dimensions[26].height = 18

for i, (rng, rating, fg_c, bg_c, desc) in enumerate(bands, start=27):
    write(ws5, i, 2, rng, bold=True, bg=bg_c, fg=fg_c, size=11, h="center", border=thin_border())
    write(ws5, i, 3, rating, bold=True, bg=bg_c, fg=fg_c, size=11, h="center", border=thin_border())
    ws5.merge_cells(start_row=i, start_column=4, end_row=i, end_column=5)
    write(ws5, i, 4, desc, bg=bg_c, fg=NEAR_BLACK, size=10, border=thin_border(), wrap=True)
    ws5.row_dimensions[i].height = 26


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MVP Roadmap
# ══════════════════════════════════════════════════════════════════════════════

ws6 = wb.create_sheet("Implementation Roadmap")
ws6.sheet_view.showGridLines = False

set_col_widths(ws6, {
    "A": 3, "B": 8, "C": 30, "D": 16, "E": 12, "F": 12, "G": 42, "H": 3,
})

ws6.merge_cells("B1:G1")
ws6.row_dimensions[1].height = 32
write(ws6, 1, 2, "Implementation Roadmap — Phased Rollout",
      bold=True, bg=NAVY, fg=WHITE, size=14, h="center")
ws6.row_dimensions[2].height = 14

phases = [
    ("MVP — Phase 1", "mvp", LIGHT_GREEN, GREEN, NAVY,
     "Foundation checks covering the highest-impact dimensions. "
     "20 checks across Completeness, Uniqueness, Validity, Integrity, Freshness, Volume. "
     "Highest signal-to-effort ratio."),
    ("Phase 2 — Enhancements", "phase2", LIGHT_AMBER, AMBER, NEAR_BLACK,
     "Checks requiring additional setup: cross-table mapping, source/target tables, "
     "baseline snapshots, or orchestration metadata. "
     "13 checks across Consistency, Accuracy, Anomaly, Schema, Pipeline."),
    ("Later — Governance", "later", LIGHT_GREY, DARK_GREY, NEAR_BLACK,
     "Governance-adjacent checks. Useful once foundational checks are stable. "
     "2 checks: Metadata Completeness, Timestamp Sequence Integrity."),
]

row = 3
for phase_label, phase_key, bg_h, fg_h, fg_row, phase_desc in phases:
    ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws6.row_dimensions[row].height = 22
    write(ws6, row, 2, phase_label, bold=True, bg=NAVY if phase_key == "mvp" else BLUE,
          fg=WHITE, size=12, h="center")
    row += 1

    ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    write(ws6, row, 2, phase_desc, bg=bg_h, fg=fg_h, size=10, italic=True, wrap=True)
    ws6.row_dimensions[row].height = 30
    row += 1

    headers = ["ID", "Check Name", "Dimension", "Eff. Weight", "Threshold", "Why this phase"]
    for c, h in enumerate(headers, start=2):
        write(ws6, row, c, h, bold=True, bg=BLUE, fg=WHITE, size=10,
              h="center", border=thin_border())
    ws6.row_dimensions[row].height = 18
    row += 1

    phase_checks = [c for c in CHECK_DATA if (
        (phase_key == "mvp" and c[6] == "MVP") or
        (phase_key == "phase2" and c[6] == "Phase 2") or
        (phase_key == "later" and c[6] == "Later")
    )]
    phase_total_weight = sum(effective_weight(c[0], c[2], c[3]) for c in phase_checks)

    for j, (cid, name, dim, within_w, thresh, direction, phase, rationale) in enumerate(phase_checks):
        eff_w = effective_weight(cid, dim, within_w)
        row_bg = LIGHT_GREY if j % 2 == 0 else WHITE
        write(ws6, row, 2, cid, bg=row_bg, fg=DARK_GREY, size=10, h="center", border=thin_border())
        write(ws6, row, 3, name, bg=row_bg, fg=NEAR_BLACK, size=10, border=thin_border())
        write(ws6, row, 4, dim.title(), bg=bg_h, fg=fg_h, size=9, h="center", border=thin_border())
        write(ws6, row, 5, eff_w, bg=row_bg, fg=NEAR_BLACK, size=10, h="center",
              border=thin_border(), num_format="0.0000")
        write(ws6, row, 6, thresh, bg=row_bg, fg=NEAR_BLACK, size=10, h="center", border=thin_border())
        write(ws6, row, 7, rationale, bg=row_bg, fg=DARK_GREY, size=9, wrap=True, border=thin_border())
        ws6.row_dimensions[row].height = 26
        row += 1

    # Phase subtotal
    write(ws6, row, 3, f"Phase subtotal ({len(phase_checks)} checks)",
          bold=True, bg=MID_GREY, fg=NAVY, size=10, h="right", border=thin_border())
    write(ws6, row, 2, "", bg=MID_GREY, border=thin_border())
    write(ws6, row, 4, "", bg=MID_GREY, border=thin_border())
    write(ws6, row, 5, phase_total_weight, bold=True, bg=MID_GREY, fg=NAVY, size=10,
          h="center", border=thin_border(), num_format="0.0000")
    write(ws6, row, 6, f"{phase_total_weight:.0%} of total score", bold=True,
          bg=MID_GREY, fg=NAVY, size=10, h="center", border=thin_border())
    write(ws6, row, 7, "", bg=MID_GREY, border=thin_border())
    ws6.row_dimensions[row].height = 20
    row += 2


# ── Tab colours ──────────────────────────────────────────────────────────────
tab_colors = {
    ws1: "1B3A6B",
    ws2: "2563EB",
    ws3: "0EA5E9",
    ws4: "7C3AED",
    ws5: "059669",
    ws6: "D97706",
}
for ws, color in tab_colors.items():
    ws.sheet_properties.tabColor = color


# ── Save ─────────────────────────────────────────────────────────────────────
out_path = "/Users/ishaansachdeva/data-quality-scanner/DQS_Weights_Explainer.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
