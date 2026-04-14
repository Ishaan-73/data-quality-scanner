from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dqs.checks import CHECK_REGISTRY
from dqs.config.defaults import (
    DIMENSION_WEIGHTS, CHECK_WITHIN_DIM_WEIGHTS,
    DEFAULT_THRESHOLDS, MVP_PHASE, effective_weight,
)

wb = Workbook()
ws = wb.active
ws.title = "DQS Checks"

# ── Styles ──────────────────────────────────────────────────────────────────
PHASE_FILL = {
    "mvp":    PatternFill("solid", fgColor="C6EFCE"),
    "phase2": PatternFill("solid", fgColor="FFEB9C"),
    "later":  PatternFill("solid", fgColor="EDEDED"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD        = Font(bold=True, size=10)
NORMAL      = Font(size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

DIM_COLORS = {
    "completeness":  "DDEEFF",
    "uniqueness":    "FFE8CC",
    "validity":      "E8F4E8",
    "consistency":   "FFF0E8",
    "integrity":     "EEE8FF",
    "freshness":     "FFF8CC",
    "volume":        "FFE8E8",
    "accuracy":      "E8FFFF",
    "anomaly":       "FFE8FF",
    "schema":        "F0F0F0",
    "metadata":      "F0F0F0",
    "pipeline":      "F0F0F0",
    "time_series":   "F0F0F0",
}

headers = [
    "Check ID", "Check Name", "Dimension", "Dim Weight",
    "Check Weight\n(within dim)", "Effective Weight", "Effective Weight %",
    "Threshold", "Phase", "What It Measures",
]

DESCRIPTIONS = {
    1:  "Null ratio in a single column",
    2:  "Avg null ratio across critical columns",
    3:  "Rows where all major fields are null/blank",
    4:  "Required field missing when a condition is true",
    5:  "Duplicate rows by primary key",
    6:  "Duplicate rows by business/natural key",
    7:  "Near-duplicate candidates (case/trim grouping)",
    8:  "Values that fail type casting",
    9:  "Values outside an approved enum/domain list",
    10: "String values failing a regex format check",
    11: "Numeric/date values outside min/max bounds",
    12: "Negative values in non-negative measures",
    13: "Cross-column or cross-table value mismatch",
    14: "Referential value inconsistency across tables",
    15: "Derived column disagrees with source columns",
    16: "Temporal ordering violation (e.g. end < start)",
    17: "Child rows with no matching parent (FK check)",
    18: "Records unlinked from any master/dimension",
    19: "Join coverage ratio falls below threshold",
    20: "Latest timestamp exceeds SLA freshness window",
    21: "Partition/batch not loaded within stale_days",
    22: "Event arrival latency exceeds threshold",
    23: "Day-over-day row count change beyond band",
    24: "Expected partition/date slice missing",
    25: "Statistical outliers vs reference distribution",
    26: "Aggregated metric deviates from benchmark",
    27: "Revenue = price × quantity identity check",
    28: "Z-score outlier detection on a numeric column",
    29: "Time-series sudden spike or drop detection",
    30: "Schema drift — unexpected column changes",
    31: "Missing or stale metadata catalog entries",
    32: "Pipeline run missing or failed",
    33: "Time-series trend or seasonality anomaly",
}

# ── Header row ───────────────────────────────────────────────────────────────
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[1].height = 36

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_idx, (check_id, check) in enumerate(sorted(CHECK_REGISTRY.items()), start=2):
    dim   = check.dimension
    dim_w = DIMENSION_WEIGHTS.get(dim, 0.0)
    chk_w = CHECK_WITHIN_DIM_WEIGHTS.get(dim, {}).get(check_id, 0.0)
    eff_w = effective_weight(check_id, dim)
    phase = MVP_PHASE.get(check_id, "later")
    thresh = DEFAULT_THRESHOLDS.get(check_id)
    thresh_str = "advisory" if thresh is None else str(thresh)

    row = [
        check_id,
        check.name,
        dim.replace("_", " ").title(),
        f"{dim_w:.0%}",
        f"{chk_w:.0%}",
        round(eff_w, 4),
        f"{eff_w:.2%}",
        thresh_str,
        phase,
        DESCRIPTIONS.get(check_id, ""),
    ]
    ws.append(row)

    dim_fill = PatternFill("solid", fgColor=DIM_COLORS.get(dim, "FFFFFF"))
    phase_fill = PHASE_FILL.get(phase, PatternFill())

    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = BORDER
        cell.font   = NORMAL
        # Phase column gets phase color, others get dim color
        if col_idx == 9:
            cell.fill = phase_fill
            cell.alignment = CENTER
        elif col_idx in (1, 4, 5, 6, 7, 8):
            cell.fill = dim_fill
            cell.alignment = CENTER
        else:
            cell.fill = dim_fill
            cell.alignment = LEFT

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [9, 32, 16, 11, 14, 14, 15, 11, 10, 48]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze header ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Dimension Weights")
ws2.append(["Dimension", "Weight", "# Checks", "Phase Breakdown"])
ws2.cell(1,1).font = ws2.cell(1,2).font = ws2.cell(1,3).font = ws2.cell(1,4).font = HEADER_FONT
for c in range(1,5):
    ws2.cell(1,c).fill = HEADER_FILL
    ws2.cell(1,c).alignment = CENTER
    ws2.cell(1,c).border = BORDER

from collections import Counter
dim_checks = {}
for check_id, check in CHECK_REGISTRY.items():
    dim_checks.setdefault(check.dimension, []).append(check_id)

for r, (dim, w) in enumerate(sorted(DIMENSION_WEIGHTS.items(), key=lambda x: -x[1]), start=2):
    ids = dim_checks.get(dim, [])
    phases = Counter(MVP_PHASE.get(i, "later") for i in ids)
    phase_str = "  ".join(f"{p}:{n}" for p, n in sorted(phases.items()))
    fill = PatternFill("solid", fgColor=DIM_COLORS.get(dim, "FFFFFF"))
    row = [dim.replace("_"," ").title(), f"{w:.0%}", len(ids), phase_str]
    ws2.append(row)
    for c in range(1,5):
        cell = ws2.cell(r, c)
        cell.fill = fill; cell.border = BORDER
        cell.font = NORMAL
        cell.alignment = CENTER if c != 1 else LEFT

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 28

wb.save("DQS_Checks_Reference.xlsx")
print("Saved DQS_Checks_Reference.xlsx")
