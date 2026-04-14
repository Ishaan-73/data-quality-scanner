from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

wb = Workbook()

# ── Shared styles ────────────────────────────────────────────────────────────
def hdr(bold=True, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
NORMAL  = Font(size=10)
BOLD10  = Font(bold=True, size=10)

FILLS = {
    "header_dark":  fill("1F4E79"),
    "header_blue":  fill("2E75B6"),
    "header_green": fill("375623"),
    "pass1":        fill("D6E4F0"),
    "pass2":        fill("D9EAD3"),
    "pass3":        fill("FFF2CC"),
    "high":         fill("FFD7D7"),
    "medium":       fill("FFF0CC"),
    "review":       fill("E8E8E8"),
    "clear":        fill("D9EAD3"),
    "block":        fill("FFD7D7"),
    "proceed":      fill("FFF0CC"),
    "synthetic":    fill("D6E4F0"),
    "gate_clear":   fill("D9EAD3"),
    "row_alt":      fill("F5F5F5"),
    "white":        fill("FFFFFF"),
}

def write_header_row(ws, row, values, col_fills, text_color="FFFFFF", row_height=28):
    for i, (val, f) in enumerate(zip(values, col_fills), 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font      = Font(bold=True, color=text_color, size=10)
        c.fill      = f
        c.alignment = CENTER
        c.border    = border()
    ws.row_dimensions[row].height = row_height

def write_row(ws, row, values, row_fill=None, aligns=None, bold_cols=None, col_fills=None):
    aligns    = aligns    or []
    bold_cols = bold_cols or []
    col_fills = col_fills or []
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font      = Font(bold=(i in bold_cols), size=10)
        c.fill      = col_fills[i-1] if col_fills and i <= len(col_fills) else (row_fill or FILLS["white"])
        c.alignment = aligns[i-1] if aligns and i <= len(aligns) else LEFT
        c.border    = border()
    ws.row_dimensions[row].height = 18

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Overview: Three Passes
# ════════════════════════════════════════════════════════════════════════════
ws1: Worksheet = wb.active
ws1.title = "How It Works"

write_header_row(ws1, 1,
    ["Pass", "Name", "Data Access", "What It Does", "Columns In Scope", "Output"],
    [FILLS["header_dark"]] * 6)

passes = [
    ("Pass 1", "Column Name Heuristics",
     "Zero queries — schema only",
     "Pattern-matches column names against a curated keyword list. No data is read.",
     "All columns in all tables referenced by the scan config.",
     "Each column tagged HIGH, MEDIUM, or CLEAR."),
    ("Pass 2", "Value Sampling + Regex",
     "1 query per flagged column (SELECT col … LIMIT 100)",
     "Pulls up to 100 non-null values per HIGH/MEDIUM column and tests them against "
     "12 PII regex patterns (email, phone, SSN, IBAN, IP, postcode, credit card, etc.).",
     "Only columns flagged HIGH or MEDIUM in Pass 1.",
     "Status upgraded to HIGH (≥80% match), kept MEDIUM (40–80%), or downgraded to REVIEW (<40%)."),
    ("Pass 3\n(opt-in)", "Semantic / NER Scan",
     "1 query per REVIEW column (SELECT col … LIMIT 100)",
     "Uses spaCy en_core_web_sm to detect named entities (PERSON, LOC, DATE, etc.) "
     "in free-text REVIEW columns. Skipped silently if spaCy is not installed.",
     "Only columns in REVIEW status after Pass 2.",
     "REVIEW → HIGH if entities found, REVIEW → CLEAR if none."),
]

pass_fills = [FILLS["pass1"], FILLS["pass2"], FILLS["pass3"]]
for r, (p, name, access, does, scope, out) in enumerate(passes, 2):
    row_fill = pass_fills[r - 2]
    write_row(ws1, r,
        [p, name, access, does, scope, out],
        row_fill=row_fill,
        aligns=[CENTER, LEFT, CENTER, LEFT, LEFT, LEFT],
        bold_cols=[1, 2])
    ws1.row_dimensions[r].height = 60

set_col_widths(ws1, [10, 24, 28, 52, 36, 46])
ws1.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Keywords (Pass 1)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pass 1 — Keywords")

write_header_row(ws2, 1,
    ["Confidence", "Keyword / Pattern", "Match Type", "Category", "PII Type"],
    [FILLS["header_dark"]] * 5)

HIGH_KEYWORDS = [
    ("HIGH", "email",          "exact", "contact",   "Contact — email"),
    ("HIGH", "ssn",            "exact", "identity",  "Identity — SSN"),
    ("HIGH", "dob",            "exact", "identity",  "Identity — date of birth"),
    ("HIGH", "date_of_birth",  "exact", "identity",  "Identity — date of birth"),
    ("HIGH", "phone",          "exact", "contact",   "Contact — phone"),
    ("HIGH", "phone_number",   "exact", "contact",   "Contact — phone"),
    ("HIGH", "first_name",     "exact", "identity",  "Identity — name"),
    ("HIGH", "last_name",      "exact", "identity",  "Identity — name"),
    ("HIGH", "full_name",      "exact", "identity",  "Identity — name"),
    ("HIGH", "passport",       "exact", "identity",  "Identity — passport"),
    ("HIGH", "national_id",    "exact", "identity",  "Identity — national ID"),
    ("HIGH", "ip_address",     "exact", "contact",   "Contact — IP address"),
    ("HIGH", "credit_card",    "exact", "financial", "Financial — credit card"),
    ("HIGH", "iban",           "exact", "financial", "Financial — IBAN"),
    ("HIGH", "sort_code",      "exact", "financial", "Financial — sort code"),
    ("HIGH", "bank_account",   "exact", "financial", "Financial — bank account"),
    ("HIGH", "tax_id",         "exact", "financial", "Financial — tax ID"),
    ("HIGH", "medical_record", "exact", "health",    "Health — medical record"),
    ("HIGH", "diagnosis",      "exact", "health",    "Health — diagnosis"),
    ("HIGH", "biometric",      "exact", "health",    "Health — biometric"),
    ("HIGH", "device_id",      "exact", "behavioral","Behavioral — device ID"),
    ("HIGH", "mac_address",    "exact", "behavioral","Behavioral — MAC address"),
    ("MEDIUM", "name",      "contains", "identity",  "Identity — name (partial)"),
    ("MEDIUM", "contact",   "contains", "contact",   "Contact (partial)"),
    ("MEDIUM", "address",   "contains", "contact",   "Contact — address"),
    ("MEDIUM", "birth",     "contains", "identity",  "Identity — birth date"),
    ("MEDIUM", "card",      "contains", "financial", "Financial — card"),
    ("MEDIUM", "account",   "contains", "financial", "Financial — account"),
    ("MEDIUM", "identity",  "contains", "identity",  "Identity"),
    ("MEDIUM", "mobile",    "contains", "contact",   "Contact — mobile"),
    ("MEDIUM", "location",  "contains", "contact",   "Contact — location"),
    ("MEDIUM", "geo",       "contains", "contact",   "Contact — geolocation"),
    ("MEDIUM", "coords",    "contains", "contact",   "Contact — coordinates"),
]

for r, (conf, kw, match, cat, pii_type) in enumerate(HIGH_KEYWORDS, 2):
    row_fill = FILLS["high"] if conf == "HIGH" else FILLS["medium"]
    conf_fill = FILLS["high"] if conf == "HIGH" else FILLS["medium"]
    write_row(ws2, r,
        [conf, kw, match, cat, pii_type],
        row_fill=row_fill,
        aligns=[CENTER, LEFT, CENTER, CENTER, LEFT],
        bold_cols=[1, 2])

set_col_widths(ws2, [12, 20, 14, 14, 32])
ws2.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Regex Patterns (Pass 2)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Pass 2 — Regex Patterns")

write_header_row(ws3, 1,
    ["Pattern Name", "Category", "PII Type", "Regex", "Example Match", "Upgrade Rule"],
    [FILLS["header_dark"]] * 6)

patterns = [
    ("email",        "contact",   "Contact — email",
     r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$",
     "user@example.com", "≥80% → HIGH"),
    ("phone_e164",   "contact",   "Contact — phone",
     r"^\+?[1-9]\d{7,14}$",
     "+447911123456", "≥80% → HIGH"),
    ("phone_us",     "contact",   "Contact — phone",
     r"^(\+1[-.\s]?)?\(?\d{3}\)?[-.\s]\d{3}[-.\s]\d{4}$",
     "(212) 555-1234", "≥80% → HIGH"),
    ("phone_uk",     "contact",   "Contact — phone",
     r"^(\+44|0)7\d{9}$",
     "07700900000", "≥80% → HIGH"),
    ("ssn_us",       "identity",  "Identity — SSN",
     r"^\d{3}-\d{2}-\d{4}$",
     "123-45-6789", "≥80% → HIGH"),
    ("ni_uk",        "identity",  "Identity — NI number",
     r"^[A-CEGHJ-PR-TW-Z]{2}\d{6}[A-D]$",
     "AB123456C", "≥80% → HIGH"),
    ("postcode_uk",  "contact",   "Contact — postcode",
     r"^[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}$",
     "SW1A 1AA", "≥80% → HIGH"),
    ("zip_us",       "contact",   "Contact — ZIP code",
     r"^\d{5}(-\d{4})?$",
     "10001", "≥80% → HIGH"),
    ("ipv4",         "contact",   "Contact — IP address",
     r"^(\d{1,3}\.){3}\d{1,3}$",
     "192.168.1.1", "≥80% → HIGH"),
    ("ipv6",         "contact",   "Contact — IP address",
     r"^([0-9a-fA-F]{0,4}:){2,7}[0-9a-fA-F]{0,4}$",
     "2001:0db8::1", "≥80% → HIGH"),
    ("credit_card",  "financial", "Financial — credit card",
     r"^\d{13,19}$",
     "4111111111111111", "≥80% → HIGH"),
    ("iban",         "financial", "Financial — IBAN",
     r"^[A-Z]{2}\d{2}[A-Z0-9]{4,30}$",
     "GB29NWBK60161331926819", "≥80% → HIGH"),
]

CAT_FILLS = {
    "contact":   fill("D6E4F0"),
    "identity":  fill("FCE4D6"),
    "financial": fill("E2EFDA"),
    "health":    fill("F4CCFF"),
    "behavioral":fill("FFF2CC"),
}

for r, (name, cat, pii, regex, example, rule) in enumerate(patterns, 2):
    row_fill = CAT_FILLS.get(cat, FILLS["white"])
    write_row(ws3, r,
        [name, cat, pii, regex, example, rule],
        row_fill=row_fill,
        aligns=[LEFT, CENTER, LEFT, LEFT, LEFT, CENTER],
        bold_cols=[1])
    ws3.row_dimensions[r].height = 20

# Confidence upgrade legend below
legend_row = len(patterns) + 3
ws3.cell(legend_row, 1, "Match Ratio Rules").font = BOLD10
ws3.cell(legend_row, 1).fill = FILLS["header_blue"]
ws3.cell(legend_row, 1).font = Font(bold=True, color="FFFFFF", size=10)
ws3.cell(legend_row, 1).alignment = CENTER
ws3.merge_cells(f"A{legend_row}:F{legend_row}")

rules = [
    ("≥ 80% of sampled values match", "→ Upgrade to HIGH",    FILLS["high"]),
    ("40–80% match",                  "→ Keep / set MEDIUM",   FILLS["medium"]),
    ("< 40% match (was MEDIUM)",      "→ Downgrade to REVIEW", FILLS["review"]),
]
for i, (cond, action, rf) in enumerate(rules, legend_row + 1):
    ws3.cell(i, 1, cond).fill  = rf
    ws3.cell(i, 1).alignment   = LEFT
    ws3.cell(i, 1).border      = border()
    ws3.cell(i, 2, action).fill = rf
    ws3.cell(i, 2).alignment   = LEFT
    ws3.cell(i, 2).border      = border()
    ws3.merge_cells(f"B{i}:F{i}")

set_col_widths(ws3, [16, 12, 26, 50, 26, 16])
ws3.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Status Values & Gate Decisions
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Status & Gate")

# ── Column Status ────────────────────────────────────────────────────────────
ws4.cell(1, 1, "Column Status Values").font = Font(bold=True, color="FFFFFF", size=11)
ws4.cell(1, 1).fill = FILLS["header_dark"]
ws4.cell(1, 1).alignment = CENTER
ws4.merge_cells("A1:E1")
ws4.row_dimensions[1].height = 24

write_header_row(ws4, 2,
    ["Status", "Meaning", "DQS Behaviour", "Detected Via", "Masked Sample Stored?"],
    [FILLS["header_blue"]] * 5, text_color="FFFFFF")

statuses = [
    ("HIGH",   "High-confidence PII column.",
     "Column excluded from all value-sampling DQS checks (check skipped, score neutral).",
     "Name match exact + regex ≥80%, or Pass 3 NER hit.", "Yes — up to 3 masked samples"),
    ("MEDIUM", "Suspected PII — not confirmed by regex.",
     "Treated same as HIGH: column excluded from value-sampling checks.",
     "Name substring match, regex 40–80%.", "Yes — up to 3 masked samples"),
    ("REVIEW", "Ambiguous — name matched but regex did not confirm.",
     "Flagged in manifest for human review. DQS checks still run.",
     "Name substring match + regex <40%.", "No"),
    ("CLEAR",  "No PII signal detected.",
     "DQS checks run normally on this column.",
     "No name or regex match.", "No"),
]
status_fills = [FILLS["high"], FILLS["medium"], FILLS["review"], FILLS["clear"]]
for r, (status, meaning, behaviour, via, masked) in enumerate(statuses, 3):
    write_row(ws4, r,
        [status, meaning, behaviour, via, masked],
        row_fill=status_fills[r - 3],
        aligns=[CENTER, LEFT, LEFT, LEFT, CENTER],
        bold_cols=[1])
    ws4.row_dimensions[r].height = 42

# ── Gate Decisions ───────────────────────────────────────────────────────────
gate_start = 9
ws4.cell(gate_start, 1, "Gate Decisions").font = Font(bold=True, color="FFFFFF", size=11)
ws4.cell(gate_start, 1).fill = FILLS["header_dark"]
ws4.cell(gate_start, 1).alignment = CENTER
ws4.merge_cells(f"A{gate_start}:E{gate_start}")
ws4.row_dimensions[gate_start].height = 24

write_header_row(ws4, gate_start + 1,
    ["Decision", "Triggered When", "on_pii_detected setting", "DQS Behaviour", "Scan Proceeds?"],
    [FILLS["header_blue"]] * 5, text_color="FFFFFF")

gates = [
    ("BLOCK",
     "HIGH column found",
     "block",
     "Scan stops immediately. PIIManifest returned. Customer must remediate.",
     "No — blocked"),
    ("SYNTHETIC_MODE",
     "HIGH column found",
     "synthetic",
     "Mode forced to B (Synthetic Clone). Profile extracted, then production disconnected. "
     "Checks run on synthetic data.",
     "Yes — synthetic only"),
    ("PROCEED_EXCLUDE",
     "HIGH or MEDIUM column found",
     "exclude (default)",
     "DQS runs. HIGH/MEDIUM columns skipped in value-sampling checks. Score unaffected.",
     "Yes — with exclusions"),
    ("CLEAR",
     "No HIGH or MEDIUM findings",
     "any",
     "DQS runs in full. Manifest attached as clean confirmation.",
     "Yes — full scan"),
]
gate_fills = [FILLS["block"], FILLS["synthetic"], FILLS["proceed"], FILLS["gate_clear"]]
for r, (dec, when, setting, behaviour, proceeds) in enumerate(gates, gate_start + 2):
    write_row(ws4, r,
        [dec, when, setting, behaviour, proceeds],
        row_fill=gate_fills[r - (gate_start + 2)],
        aligns=[CENTER, LEFT, CENTER, LEFT, CENTER],
        bold_cols=[1])
    ws4.row_dimensions[r].height = 48

set_col_widths(ws4, [20, 28, 22, 52, 16])
ws4.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Masking Rules
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Masking Rules")

write_header_row(ws5, 1,
    ["Category", "Rule", "Raw Example", "Masked Output"],
    [FILLS["header_dark"]] * 4)

masking = [
    ("contact (email)",   "Preserve domain; mask local part after first char.",
     "john.doe@company.com", "j***@company.com"),
    ("contact (phone)",   "Preserve first 3 chars (country code); mask digits.",
     "+447911123456", "+44***********"),
    ("identity (name)",   "Preserve first char of each token.",
     "John Smith", "J*** S***"),
    ("identity (date)",   "Preserve century; mask rest.",
     "1985-03-22", "19**-**-**"),
    ("financial (card)",  "Preserve last 4 digits only.",
     "4111111111111111", "**** **** **** 1111"),
    ("other",             "Preserve first 2 chars; mask remainder.",
     "AB123456C",  "AB*******"),
]

for r, (cat, rule, raw, masked) in enumerate(masking, 2):
    rf = fill("F5F5F5") if r % 2 == 0 else FILLS["white"]
    write_row(ws5, r, [cat, rule, raw, masked], row_fill=rf,
        aligns=[LEFT, LEFT, LEFT, LEFT], bold_cols=[1])
    ws5.row_dimensions[r].height = 22

# Note
note_row = len(masking) + 3
ws5.cell(note_row, 1,
    "Raw values are NEVER stored. Only masked samples (≤3 per column) are written to the manifest."
).font = Font(bold=True, italic=True, size=10, color="C00000")
ws5.merge_cells(f"A{note_row}:D{note_row}")
ws5.cell(note_row, 1).alignment = LEFT

set_col_widths(ws5, [22, 48, 26, 26])
ws5.freeze_panes = "A2"


wb.save("DQS_PII_Screener_Reference.xlsx")
print("Saved DQS_PII_Screener_Reference.xlsx")
